# coding:utf-8
from openpyxl import *

class OperationExcel:
    def __init__(self,file_name=None,sheet_id=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id
            self.wb = load_workbook(self.file_name)
            sheets = self.wb.sheetnames
            self.sheet = sheets[self.sheet_id]
            self.tables = self.wb[self.sheet]
        else:
            self.file_name = '../dataconfig/interfaceTest.xlsx'
            self.sheet_id = 0
            self.wb = load_workbook(self.file_name)
            sheets = self.wb.sheetnames
            self.sheet = sheets[self.sheet_id]
            self.tables = self.wb[self.sheet]

    # 获取sheet
    def get_data(self):
        sheet_name = self.wb.sheetnames[self.sheet_id]
        return sheet_name

    def get_test(self):
        pass

    # 获取workbook所有sheet名
    def get_sheet_list(self):
        sheets_list = self.wb.sheetnames
        return sheets_list

    # 获取表内最大有效行数
    def get_lines(self):
        rows = self.tables.max_row
        return rows

    # 根据行、列获取单元格
    def get_cell_value(self,row,col):
        cellvalue = self.tables.cell(row=row, column=col).value
        return cellvalue

    # 在单元格内写入值
    def write_value(self, row, colunm, cellvalue):
        try:
            self.tables.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file_name)
            return 'Write Successfully!'
        except:
            self.tables.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file_name)

    # 根据对应的caseid找到对应行的内容
    def get_rows_data(self,case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    # 根据对应的caseid找到对应的行号
    def get_row_num(self,case_id):
        num = 0
        cols_data = self.get_cols_data()
        for col_data in cols_data:
            if case_id in col_data:
                return num
            num = num+1

    # 根据行号找到该行的内容
    def get_row_values(self,row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data

    # 获取某一列的内容
    def get_cols_data(self,col_id=None):
        if col_id != None:
            cols = self.data.col_values(col_id)
        else:
            cols = self.data.col_values(0)
        return cols

if __name__ == '__main__':
    file_name = '../dataconfig/interfaceTest.xlsx'
    sheet_id = 0
    testexcel = OperationExcel(file_name,sheet_id)
    print(testexcel.get_lines())
    print(testexcel.get_sheet_list())
    print(testexcel.get_cell_value(2,1))
    print(testexcel.get_data())
    print("There are write call value : ",testexcel.write_value(11,1,'白妖大大宇宙第一无敌帅'))
    print(testexcel.get_test())